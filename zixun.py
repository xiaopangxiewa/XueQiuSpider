# -*- coding: utf-8 -*-
import scrapy
from  xueqiu.items import XueqiuItem
import re,time
from scrapy.http import Request,FormRequest
import urllib.request
from openpyxl import load_workbook



class ZixunSpider(scrapy.Spider):
    name = "zixun"
    allowed_domains = ["xueqiu.com"]
    #start_urls = ['http://xueqiu.com/']

    def start_requests(self):
        s_url='https://xueqiu.com'    #必须先爬取一下首页，将首页缓存至cookiejar中，直接爬取内容页会报错；
        yield Request(s_url,callback=self.parse,dont_filter=True,
                      meta={"cookiejar":1})
        #发送cookie请求
    

    def parse(self, response):
        print('网页爬取成功')
        a=response.body.decode('utf-8')
        print(len(a))
        d_url='https://xueqiu.com/v4/statuses/public_timeline_by_category.json?since_id=-1&max_id=-1&count=10&category=-1'
        yield Request(d_url,callback=self.next1,
                      meta={"cookiejar":response.meta["cookiejar"]})
        
        
    def next1(self,response):
        print('内容页爬取成功')
        a=response.body.decode('utf-8')
        print(len(a))
        pat_next_page_id='"next_max_id":(.*?),'
        next_page_id=re.compile(pat_next_page_id).findall(a)

        item=XueqiuItem()
        t_pat='."title.":."(.*?).",'
        titlelist=re.compile(t_pat).findall(a)
        print(len(titlelist))
        item['title']=titlelist
        d_pat='."title.":.".*?.",."description.":."(.*?)."'
        detail_list=re.compile(d_pat).findall(a)
        print(len(detail_list))
        item['detail']=detail_list
        
        wb=load_workbook(r'F:\Python\scrapy\xueqiu\xuqiuzixun.xlsx')
        st = wb.get_sheet_by_name("Sheet1")
        for i in range(1,len(titlelist)+1):
            d=st.max_row
            #得到当前表的最大行数，然后从最后开始写，否则会覆盖已写入数据
            c=st.cell(column=1,row=d+1)
            d=st.cell(column=2,row=d+1)
            c.value=titlelist[i-1]
            d.value=detail_list[i-1]
        wb.save('xuqiuzixun.xlsx')
        time.sleep(3)
        yield item
        
        
    
        n_url='https://xueqiu.com/v4/statuses/public_timeline_by_category.json?since_id=-1&max_id=%s&count=10&category=-1'%(next_page_id[0])
        print(n_url)
        
            
        yield Request(n_url,callback=self.next1,meta={"cookiejar":True})

#当出现：[scrapy.spidermiddlewares.httperror] INFO: Ignoring response
# <400：HTTP status code is not handled or not allowed
#说明当前状态码没有被正确识别，在setting里面进行修改即可

    